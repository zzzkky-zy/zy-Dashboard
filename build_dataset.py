"""
一次性脚本: xlsx → CSV + 图片资产
把 /Users/yaman/Desktop/产品调研v3.2.xlsx 拍平成 data/products.csv，
并把嵌入的包装图导出到 data/images/ (按 sheet+row hash 命名,
csv 中保留 包装图文件名 列以便看板查找)。

运行:
    python3 build_dataset.py
"""
from __future__ import annotations
import hashlib
import os
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

XLSX = os.environ.get("PRODUCT_XLSX", "/Users/yaman/Desktop/产品调研v3.2.xlsx")
OUT_DIR = Path(__file__).parent / "data"
IMG_DIR = OUT_DIR / "images"
CSV_PATH = OUT_DIR / "products.csv"
TEMPLATE_PATH = OUT_DIR / "products_template.csv"

# 与 data_loader.py 保持一致的列重命名
RENAME = {
    "天猫售量\n（累计）": "天猫销量原始",
    "抖音售量\n(累计）": "抖音销量原始",
    "规格/ml": "规格",
    "单ml": "天猫单ml",
    "单ml.1": "抖音单ml",
}
PACKAGE_COL = 3  # D 列 (包装图)


def extract_images(xlsx: str, out_dir: Path) -> dict[tuple[str, int], str]:
    """提取嵌入图片到 out_dir, 返回 {(sheet, excel_row_1based): filename}."""
    out_dir.mkdir(parents=True, exist_ok=True)
    # 清空旧图
    for f in out_dir.glob("*"): f.unlink()
    mapping: dict[tuple[str, int], str] = {}
    wb = load_workbook(xlsx, data_only=True)
    for sh in wb.sheetnames:
        if sh == "总": continue
        ws = wb[sh]
        for img in getattr(ws, "_images", []):
            anchor = img.anchor
            f = getattr(anchor, "_from", None)
            if f is None: continue
            if f.col != PACKAGE_COL: continue
            try: data = img._data()
            except Exception: continue
            if not data: continue
            ext = "jpg" if data[:2] == b"\xff\xd8" else "png"
            excel_row = f.row + 1
            digest = hashlib.md5(bytes(data)).hexdigest()[:10]
            fname = f"{sh}_{excel_row}_{digest}.{ext}"
            (out_dir / fname).write_bytes(bytes(data))
            mapping[(sh, excel_row)] = fname
    return mapping


def build_csv():
    OUT_DIR.mkdir(exist_ok=True)
    print(f"reading {XLSX}")
    images = extract_images(XLSX, IMG_DIR)
    print(f"  -> exported {len(images)} package images to {IMG_DIR}")

    xls = pd.ExcelFile(XLSX)
    sheets = [s for s in xls.sheet_names if s != "总"]
    frames = []
    for sh in sheets:
        df = pd.read_excel(xls, sheet_name=sh).rename(columns=RENAME)
        df["品类"] = sh
        df["_excel_row"] = df.index + 2
        df["包装图文件名"] = df.apply(
            lambda r: images.get((sh, int(r["_excel_row"])), ""), axis=1
        )
        df = df.drop(columns=["_excel_row", "包装图"], errors="ignore")
        frames.append(df)
    full = pd.concat(frames, ignore_index=True)

    # 列序: 关键列前置, 便于他人编辑
    front = ["品类", "品牌", "产品", "规格", "爆款香型",
             "天猫挂价", "天猫单ml", "配赠", "天猫销量原始",
             "达播价", "抖音单ml", "配赠.1", "配赠（可带上机制主图）", "抖音销量原始",
             "功效词", "一句话卖点", "详情页宣传（截图）", "小红书用户评价",
             "包装图文件名"]
    cols = [c for c in front if c in full.columns] + \
           [c for c in full.columns if c not in front]
    full = full[cols]

    full.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")
    print(f"  -> wrote {CSV_PATH} ({len(full)} rows, {len(cols)} cols)")

    # 模板: 表头 + 几行示范
    template = full.head(2).copy()
    template.iloc[:, :] = ""
    template.to_csv(TEMPLATE_PATH, index=False, encoding="utf-8-sig")
    print(f"  -> wrote {TEMPLATE_PATH} (空白模板, 供他人补充使用)")


if __name__ == "__main__":
    build_csv()
