"""
从 xlsx 中抽取嵌入图片（包装图列），按 (sheet, row) 写入本地缓存目录。
"""
from __future__ import annotations
import hashlib
import os
from pathlib import Path

from openpyxl import load_workbook

CACHE_DIR = Path(__file__).parent / "_image_cache"

# 包装图位于 D 列 -> col 索引 3 (0-based)
PACKAGE_COL = 3


def extract_package_images(xlsx_path: str, force: bool = False) -> dict[tuple[str, int], str]:
    """
    返回 {(sheet_name, excel_row_1based): cached_png_path}
    excel_row_1based 与 pandas read_excel 的 0-based row 满足: pandas_row = excel_row - 2
    (减2: 1 表头, 1 1-based)
    """
    CACHE_DIR.mkdir(exist_ok=True)
    sig_path = CACHE_DIR / "_signature.txt"
    sig = f"{os.path.getmtime(xlsx_path)}-{os.path.getsize(xlsx_path)}"
    if sig_path.exists() and sig_path.read_text() == sig and not force:
        return _scan_existing()

    # 失效, 重建
    for f in CACHE_DIR.glob("*.png"):
        try: f.unlink()
        except OSError: pass
    for f in CACHE_DIR.glob("*.jpg"):
        try: f.unlink()
        except OSError: pass

    mapping: dict[tuple[str, int], str] = {}
    wb = load_workbook(xlsx_path, data_only=True)
    for sheet in wb.sheetnames:
        if sheet == "总":
            continue
        ws = wb[sheet]
        for img in getattr(ws, "_images", []):
            anchor = img.anchor
            f = getattr(anchor, "_from", None)
            if f is None:
                continue
            col0 = f.col  # 0-based
            row0 = f.row  # 0-based
            if col0 != PACKAGE_COL:
                continue
            try:
                data = img._data()
            except Exception:
                continue
            if not data:
                continue
            ext = "png"
            head = bytes(data[:4])
            if head.startswith(b"\xff\xd8"):
                ext = "jpg"
            elif head.startswith(b"\x89PNG"):
                ext = "png"
            else:
                ext = "img"
            excel_row = row0 + 1  # openpyxl row is 0-based -> excel 1-based
            digest = hashlib.md5(data).hexdigest()[:8]
            fname = f"{sheet}_{excel_row}_{digest}.{ext}"
            fpath = CACHE_DIR / fname
            fpath.write_bytes(bytes(data))
            mapping[(sheet, excel_row)] = str(fpath)

    sig_path.write_text(sig)
    # 也持久化映射, 方便再次冷启动
    idx = CACHE_DIR / "_index.tsv"
    with idx.open("w") as fp:
        for (s, r), p in mapping.items():
            fp.write(f"{s}\t{r}\t{p}\n")
    return mapping


def _scan_existing() -> dict[tuple[str, int], str]:
    idx = CACHE_DIR / "_index.tsv"
    out: dict[tuple[str, int], str] = {}
    if not idx.exists():
        return out
    for line in idx.read_text().splitlines():
        parts = line.split("\t")
        if len(parts) == 3 and os.path.exists(parts[2]):
            out[(parts[0], int(parts[1]))] = parts[2]
    return out


if __name__ == "__main__":
    m = extract_package_images("/Users/yaman/Desktop/产品调研v3.2.xlsx", force=True)
    print(f"extracted {len(m)} images, sample:")
    for k, v in list(m.items())[:5]:
        print(" ", k, "->", v)
